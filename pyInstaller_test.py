# coding:utf-8
from selenium import webdriver

driver = webdriver.Chrome()
#driver.get("https://cn.bing.com/")

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
 
# 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
wb = Workbook()
 
#获取当前活跃的worksheet,默认就是第一个worksheet
ws = wb.active
 
#设置单元格的值，A1等于6(测试可知openpyxl的行和列编号从1开始计算)，B1等于7
ws.cell(row=1, column=1).value = 6
ws.cell(row=2, column=2).value = 7
 
#从第2行开始，写入9行10列数据，值为对应的列序号A、B、C、D...
for row in range(2,11):
    for col in range (1,11):
        ws.cell(row=row, column=col).value = get_column_letter(col)


 
#可以使用append插入一行数据
ws.append(["企业账号ID","用户名称","用户ID","用户ak","用户sk"])
for row in range(12,20):
    for col in range(1,6):
        ws.cell(row=row, column=col).value = get_column_letter(col)
 
#保存
wb.save(filename="pyinstall.xlsx")


with open("testa.xlsx",'wb+') as fp:
    for i in range(1,10):
        fp.write("test",)
        
    fp.close()